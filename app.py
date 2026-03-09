import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
import io
from datetime import datetime
import os

# ── Реєструємо шрифт з підтримкою кирилиці ──────────────────────────
# Шукаємо FreeSans у кількох можливих шляхах (локально і на Streamlit Cloud)
_FONT_SEARCH = [
    ("/usr/share/fonts/truetype/freefont/FreeSans.ttf",
     "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
     "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
]

_FONTS_OK = False
PDF_FONT      = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"

for _reg, _bold in _FONT_SEARCH:
    if os.path.exists(_reg) and os.path.exists(_bold):
        try:
            pdfmetrics.registerFont(TTFont("CyrillicRegular", _reg))
            pdfmetrics.registerFont(TTFont("CyrillicBold",    _bold))
            PDF_FONT      = "CyrillicRegular"
            PDF_FONT_BOLD = "CyrillicBold"
            _FONTS_OK = True
            break
        except Exception:
            continue

# Якщо жоден шрифт не знайдено — шукаємо будь-який ttf з підтримкою Unicode
if not _FONTS_OK:
    import glob
    for _ttf in glob.glob("/usr/share/fonts/**/*.ttf", recursive=True):
        try:
            pdfmetrics.registerFont(TTFont("CyrillicRegular", _ttf))
            PDF_FONT      = "CyrillicRegular"
            PDF_FONT_BOLD = "CyrillicRegular"
            _FONTS_OK = True
            break
        except Exception:
            continue

# ─────────────────────────────────────────────
# НАЛАШТУВАННЯ СТОРІНКИ
# ─────────────────────────────────────────────
st.set_page_config(page_title="Зведення щеплень", page_icon="💉", layout="wide")
st.title("💉 Зведення звітів про виконання щеплень")
st.markdown("Завантажте файли Excel від ЗОЗ — програма перевірить їх і зведе в один звіт.")
st.divider()


# ─────────────────────────────────────────────
# ДОПОМІЖНІ ФУНКЦІЇ
# ─────────────────────────────────────────────

def safe_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    try: return float(str(val).replace(",", ".").strip())
    except: return 0


def get_col_letter(col_idx):
    result = ""
    while col_idx > 0:
        col_idx, r = divmod(col_idx - 1, 26)
        result = chr(65 + r) + result
    return result


def validate_file(file_bytes, filename):
    """Повна валідація одного файлу. Повертає dict з результатами."""
    errors, warnings = [], []
    name, edrpou, period, wb = "—", "—", "—", None
    fixable = {}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        required = ["План", "Виконання", "Залишки", "Зведений звіт", "Аркуш1"]
        missing = [s for s in required if s not in wb.sheetnames]
        if missing:
            errors.append(f"Відсутні аркуші: {', '.join(missing)}")
            return dict(file=filename, name=name, edrpou=edrpou, period=period,
                        status="🔴 Помилки", errors=errors, warnings=warnings,
                        wb=wb, fixable=fixable)

        ws_plan = wb["План"]
        ws_exec = wb["Виконання"]
        ws_rem  = wb["Залишки"]
        ws_zvit = wb["Зведений звіт"]

        # ── Назва закладу ─────────────────────────────────────────────
        # Первинне джерело: Виконання!C4 (заповнене у всіх реальних файлах)
        # Резервне: План!D8
        raw_name = ws_exec["C4"].value or ws_plan["D8"].value
        name = str(raw_name).strip() if raw_name else "—"
        if not raw_name or name == "":
            errors.append("Порожня назва закладу (Виконання!C4)")
            fixable["org_name"] = ""
        else:
            fixable["org_name_current"] = name

        # ── ЄДРПОУ ───────────────────────────────────────────────────
        # Первинне джерело: Виконання!F4
        # Резервне: План!E8
        # Примітка: ЄДРПОУ в рядках даних аркушу "План" (col B) —
        # завжди містить код ЦКПХ (шаблон), не перевіряємо
        raw_edrpou = ws_exec["F4"].value or ws_plan["E8"].value
        edrpou = str(raw_edrpou).strip().lstrip("'") if raw_edrpou else "—"
        if not raw_edrpou:
            errors.append("Порожній код ЄДРПОУ (Виконання!F4)")
            fixable["org_edrpou"] = ""
        elif not str(edrpou).isdigit() or len(str(edrpou)) not in (7, 8, 9, 10):
            warnings.append(f"Код ЄДРПОУ '{edrpou}' має нестандартну довжину або містить не цифри")
            fixable["org_edrpou"] = edrpou
        else:
            fixable["org_edrpou_current"] = edrpou

        # ── Звітний період ───────────────────────────────────────────
        raw_period = ws_exec["F6"].value
        period = str(raw_period) if raw_period else "—"
        if not raw_period:
            errors.append("Відсутній звітний період (Виконання!F6)")
            fixable["report_period"] = True

        # ── Узгодженість назви між аркушами ─────────────────────────
        names_all = {
            "Виконання!C4":  str(ws_exec["C4"].value or "").strip(),
            "Залишки!A3":    str(ws_rem.cell(3, 1).value or "").strip(),
            "Зведений!A3":   str(ws_zvit.cell(3, 1).value or "").strip(),
        }
        non_empty_names = [v for v in names_all.values() if v]
        if len(set(non_empty_names)) > 1:
            warnings.append(f"Назва закладу різниться між аркушами: {names_all}")

        # ── Від'ємні значення у Виконанні ────────────────────────────
        neg_values = {}
        for row in range(8, 105):
            val = ws_exec.cell(row=row, column=5).value
            if isinstance(val, (int, float)) and val < 0:
                vac = ws_exec.cell(row=row, column=3).value
                age = ws_exec.cell(row=row, column=4).value
                errors.append(f"Від'ємна кількість щеплень: {vac} / {age} = {val}")
                neg_values[row] = {"vaccine": str(vac or ""), "age": str(age or ""), "value": val}
        if neg_values:
            fixable["neg_values"] = neg_values

        # ── Балансова формула залишків ────────────────────────────────
        balance_errors = {}
        for row in range(11, 38):
            vaccine = ws_rem.cell(row=row, column=1).value
            if not vaccine: continue
            b     = safe_num(ws_rem.cell(row=row, column=2).value)
            c     = safe_num(ws_rem.cell(row=row, column=3).value)
            d     = ws_rem.cell(row=row, column=4).value
            f_val = safe_num(ws_rem.cell(row=row, column=6).value)
            g     = safe_num(ws_rem.cell(row=row, column=7).value)
            h     = safe_num(ws_rem.cell(row=row, column=8).value)
            expected = b + c + g + h - f_val
            if isinstance(d, (int, float)):
                if abs(d - expected) > 0.5:
                    errors.append(f"Залишки — помилка балансу для «{str(vaccine).strip()}» (є {d}, має бути {expected})")
                    balance_errors[row] = {"vaccine": str(vaccine).strip(), "current": d, "expected": expected}
                if d < 0:
                    errors.append(f"Залишки — від'ємний залишок для «{str(vaccine).strip()}»")
        if balance_errors:
            fixable["balance_errors"] = balance_errors

        # ── НОВА: Залишки — Використано ≥ Виконано ───────────────────
        for row in range(11, 38):
            vaccine = ws_rem.cell(row=row, column=1).value
            if not vaccine: continue
            used = safe_num(ws_rem.cell(row=row, column=6).value)  # col F
            done = safe_num(ws_rem.cell(row=row, column=5).value)  # col E
            if done > 0 and used < done:
                warnings.append(
                    f"Залишки «{str(vaccine).strip()}»: використано ({int(used)}) < "
                    f"виконано щеплень ({int(done)}) — фізично неможливо"
                )

        # ── Протипокази ───────────────────────────────────────────────
        contra_errors = {}
        for row in range(8, 11):
            p = safe_num(ws_exec.cell(row=row, column=16).value)
            q = safe_num(ws_exec.cell(row=row, column=17).value)
            r = ws_exec.cell(row=row, column=18).value
            if isinstance(r, (int, float)) and abs(r - (p + q)) > 0.5:
                errors.append(f"Рядок {row}: протипокази ВСЬОГО ({r}) ≠ Тимчасові+Постійні ({p+q})")
                contra_errors[row] = {"temp": p, "perm": q, "total": r}
        if contra_errors:
            fixable["contra_errors"] = contra_errors

        # ── КДП-3 ─────────────────────────────────────────────────────
        for row in range(8, 11):
            l_val = safe_num(ws_exec.cell(row=row, column=12).value)
            m_val = safe_num(ws_exec.cell(row=row, column=13).value)
            if m_val > l_val > 0:
                errors.append(f"Рядок {row}: «Отримали КДП-3» ({int(m_val)}) > «Народилося за 7 міс.» ({int(l_val)})")

        # ── НОВА: Виконання "всього" (col G) ↔ Зведений звіт "місяць" (col D) ──
        EXEC_ZVIT_PAIRS = [
            (11,  15,  "БЦЖ"),
            (23,  28,  "Поліомієліт"),
            (35,  41,  "Гепатит В"),
            (42,  49,  "КПК"),
            (48,  56,  "Hib"),
            (61,  70,  "ВПЛ"),
            (99,  114, "АКДП"),
            (100, 115, "АаКДП"),
            (101, 116, "АДП"),
            (103, 118, "АДПм"),
            (104, 119, "АП"),
        ]
        for exec_row, zvit_row, label in EXEC_ZVIT_PAIRS:
            exec_val = safe_num(ws_exec.cell(row=exec_row, column=7).value)
            zvit_val = safe_num(ws_zvit.cell(row=zvit_row, column=4).value)
            if abs(exec_val - zvit_val) > 0.5 and (exec_val > 0 or zvit_val > 0):
                errors.append(
                    f"Розбіжність «{label}»: Виконання р{exec_row} всього={int(exec_val)} ≠ "
                    f"Зведений звіт р{zvit_row} місяць={int(zvit_val)}"
                )

        # ── НОВА: % у Зведеному звіті = ytd / план × 100 ─────────────
        for row in range(11, 120):
            plan_v = ws_zvit.cell(row=row, column=3).value
            ytd_v  = ws_zvit.cell(row=row, column=5).value
            pct_v  = ws_zvit.cell(row=row, column=6).value
            vac_v  = ws_zvit.cell(row=row, column=1).value
            if (isinstance(plan_v, (int, float)) and plan_v > 0
                    and isinstance(ytd_v, (int, float))
                    and isinstance(pct_v, (int, float))):
                expected_pct = round(ytd_v / plan_v * 100, 2)
                if abs(pct_v - expected_pct) > 1.0:
                    warnings.append(
                        f"Зведений звіт р{row} «{str(vac_v or '').strip()}»: "
                        f"% у файлі={round(pct_v, 1)}, розрахунковий={expected_pct} "
                        f"(план={int(plan_v)}, ytd={int(ytd_v)})"
                    )

        # ── НОВА: План (col F) ↔ Зведений звіт "річний план" (col C) ─
        PLAN_ZVIT_PLAN_PAIRS = [
            (12,  34,  "Геп В 3, до 1 р"),
            (14,  20,  "Поліо 3, до 1 р"),
            (15,  23,  "Поліо 4, 18 міс"),
            (16,  26,  "Поліо 5, 6 р"),
            (17,  81,  "АКДП-3, до 1 р"),
            (20,  100, "АДПм, 16 р"),
            (21,  108, "АДПм рев, дорослі"),
            (24,  43,  "КПК-1, 1 рік"),
            (25,  46,  "КПК-2, 4 роки"),
            (26,  57,  "ВПЛ 1 доза, 12 р"),
            (27,  58,  "ВПЛ 1 доза, 13 р"),
        ]
        for plan_row, zvit_row, label in PLAN_ZVIT_PLAN_PAIRS:
            plan_v = safe_num(ws_plan.cell(row=plan_row, column=6).value)
            zvit_v = safe_num(ws_zvit.cell(row=zvit_row, column=3).value)
            if abs(plan_v - zvit_v) > 0.5 and (plan_v > 0 or zvit_v > 0):
                warnings.append(
                    f"Розбіжність планів «{label}»: "
                    f"План р{plan_row}={int(plan_v)} ≠ Зведений р{zvit_row}={int(zvit_v)}"
                )

    except Exception as e:
        errors.append(f"Не вдалось прочитати файл: {e}")

    status = "🔴 Помилки" if errors else ("🟡 Попередження" if warnings else "🟢 OK")
    return dict(file=filename, name=name, edrpou=edrpou, period=period,
                status=status, errors=errors, warnings=warnings, wb=wb, fixable=fixable)


def apply_corrections(file_bytes, corr):
    """Застосовує виправлення до файлу. Повертає нові байти."""
    wb = load_workbook(io.BytesIO(file_bytes))

    # Назва закладу
    if corr.get("org_name"):
        v = corr["org_name"]
        wb["План"]["D8"].value          = v
        wb["Виконання"]["C4"].value     = v
        wb["Залишки"]["A4"].value       = v
        wb["Зведений звіт"]["A3"].value = v

    # ЄДРПОУ
    if corr.get("org_edrpou"):
        v = corr["org_edrpou"]
        wb["План"]["E8"].value          = v
        wb["Виконання"]["F4"].value     = v
        wb["Залишки"]["D4"].value       = v
        wb["Зведений звіт"]["D3"].value = v

    # Звітний період
    if corr.get("report_period"):
        v = corr["report_period"]
        wb["Виконання"]["F6"].value = v
        wb["Залишки"]["D6"].value   = v
        wb["Зведений звіт"]["D5"].value = v

    # Від'ємні значення
    for row_str, val in corr.get("neg_values", {}).items():
        wb["Виконання"].cell(row=int(row_str), column=5).value = max(0, val)

    # Залишки (перезаписуємо closing = b+c+g+h-f)
    for row_str, _ in corr.get("balance_accept", {}).items():
        row = int(row_str)
        ws = wb["Залишки"]
        b = safe_num(ws.cell(row=row, column=2).value)
        c = safe_num(ws.cell(row=row, column=3).value)
        f = safe_num(ws.cell(row=row, column=6).value)
        g = safe_num(ws.cell(row=row, column=7).value)
        h = safe_num(ws.cell(row=row, column=8).value)
        ws.cell(row=row, column=4).value = b + c + g + h - f

    # Протипокази (ВСЬОГО = Тимчасові + Постійні)
    for row_str, vals in corr.get("contra_fix", {}).items():
        row = int(row_str)
        t = vals.get("temp", 0)
        p = vals.get("perm", 0)
        wb["Виконання"].cell(row=row, column=16).value = t
        wb["Виконання"].cell(row=row, column=17).value = p
        wb["Виконання"].cell(row=row, column=18).value = t + p

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_pdf_report(results, corrections_log, org_name, report_label):
    """Генерує PDF-звіт про перевірку файлів."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    # Стилі
    def ps(name, size=10, bold=False, color=colors.black, indent=0):
        return ParagraphStyle(name, fontName=PDF_FONT_BOLD if bold else PDF_FONT,
                              fontSize=size, textColor=color,
                              leftIndent=indent, spaceAfter=2)

    style_title    = ps("title",    16, bold=True)
    style_subtitle = ps("subtitle", 11, bold=True, color=colors.HexColor("#1F4E79"))
    style_h2       = ps("h2",       10, bold=True, color=colors.HexColor("#2E75B6"))
    style_normal   = ps("normal",   9)
    style_error    = ps("error",    9,  color=colors.HexColor("#C00000"), indent=10)
    style_warning  = ps("warning",  9,  color=colors.HexColor("#C55A00"), indent=10)
    style_ok       = ps("ok",       9,  color=colors.HexColor("#375623"), indent=10)
    style_fix      = ps("fix",      9,  color=colors.HexColor("#7030A0"), indent=10)
    style_small    = ps("small",    8,  color=colors.grey)

    story = []

    # Заголовок
    story.append(Paragraph("Звіт про перевірку файлів щеплень", style_title))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"Організація: {org_name}", style_subtitle))
    story.append(Paragraph(f"Звітний період: {report_label}", style_normal))
    story.append(Paragraph(f"Дата формування: {datetime.now().strftime('%d.%m.%Y %H:%M')}", style_small))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1F4E79")))
    story.append(Spacer(1, 0.3*cm))

    # Зведена таблиця
    ok_count   = sum(1 for r in results if r["status"] == "🟢 OK")
    warn_count = sum(1 for r in results if r["status"] == "🟡 Попередження")
    err_count  = sum(1 for r in results if r["status"] == "🔴 Помилки")
    corr_count = len(corrections_log)

    story.append(Paragraph("Зведена статистика", style_h2))
    summary_data = [
        ["Показник", "Кількість"],
        ["Всього файлів",       str(len(results))],
        ["Без помилок (OK)",    str(ok_count)],
        ["З попередженнями",    str(warn_count)],
        ["З критичними помилками", str(err_count)],
        ["Виправлено онлайн",   str(corr_count)],
    ]
    t = Table(summary_data, colWidths=[10*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), PDF_FONT_BOLD),
        ("FONTNAME",    (0,1), (-1,-1), PDF_FONT),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#BFBFBF")),
        ("ALIGN",       (1,0), (1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWHEIGHT",   (0,0), (-1,-1), 0.6*cm),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Деталі по кожному файлу
    story.append(Paragraph("Детальні результати по файлах", style_h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#BFBFBF")))
    story.append(Spacer(1, 0.2*cm))

    for r in results:
        # Заголовок файлу
        status_label = r["status"].replace("🟢","[OK]").replace("🟡","[ПОПЕРЕДЖЕННЯ]").replace("🔴","[ПОМИЛКИ]")
        story.append(Paragraph(f"{status_label}  {r['file']}", style_subtitle))
        story.append(Paragraph(f"Заклад: {r['name']}   |   ЄДРПОУ: {r['edrpou']}   |   Період: {r['period']}", style_normal))

        if r["errors"]:
            story.append(Paragraph("Критичні помилки:", ps("eh", 9, bold=True, color=colors.HexColor("#C00000"))))
            for e in r["errors"]:
                story.append(Paragraph(f"• {e}", style_error))

        if r["warnings"]:
            story.append(Paragraph("Попередження:", ps("wh", 9, bold=True, color=colors.HexColor("#C55A00"))))
            for w in r["warnings"]:
                story.append(Paragraph(f"• {w}", style_warning))

        if not r["errors"] and not r["warnings"]:
            story.append(Paragraph("✓ Файл пройшов усі перевірки без зауважень.", style_ok))

        # Виправлення
        fname = r["file"]
        if fname in corrections_log:
            story.append(Paragraph("Виправлення внесені онлайн:", ps("ch", 9, bold=True, color=colors.HexColor("#7030A0"))))
            for fix in corrections_log[fname]:
                story.append(Paragraph(f"• {fix}", style_fix))

        story.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor("#EEEEEE")))
        story.append(Spacer(1, 0.2*cm))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def extract_facility_data(file_bytes, name, edrpou):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws_exec = wb["Виконання"]
    ws_rem  = wb["Залишки"]
    ws_zvit = wb["Зведений звіт"]

    coverage = []
    for row in range(11, 120):
        vaccine  = ws_zvit.cell(row=row, column=1).value
        age      = ws_zvit.cell(row=row, column=2).value
        plan     = ws_zvit.cell(row=row, column=3).value
        executed = ws_zvit.cell(row=row, column=4).value
        pct      = ws_zvit.cell(row=row, column=6).value
        if vaccine and isinstance(plan, (int, float)) and plan > 0:
            label = str(vaccine).strip()
            if age: label += f" ({str(age).strip()})"
            coverage.append({"label": label, "vaccine": str(vaccine).strip(),
                             "age": str(age or "").strip(), "plan": safe_num(plan),
                             "executed": safe_num(executed), "pct": safe_num(pct)})

    stocks = []
    for row in range(11, 38):
        vaccine = ws_rem.cell(row=row, column=1).value
        if not vaccine: continue
        stocks.append({"vaccine": str(vaccine).strip(),
                        "closing":  safe_num(ws_rem.cell(row=row, column=4).value),
                        "used":     safe_num(ws_rem.cell(row=row, column=6).value),
                        "opening":  safe_num(ws_rem.cell(row=row, column=2).value),
                        "received": safe_num(ws_rem.cell(row=row, column=3).value)})

    refusal_map = {8:"Туберкульоз", 9:"Поліомієліт", 10:"Гепатит В",
                   11:"Кашлюк, дифтерія, правець", 12:"Гемофільна інфекція",
                   13:"Кір, паротит, краснуха"}
    refusals = [{"disease": d, "count": safe_num(ws_exec.cell(row=r, column=20).value)}
                for r, d in refusal_map.items()]

    temp = sum(safe_num(ws_exec.cell(row=r, column=16).value) for r in range(8, 11))
    perm = sum(safe_num(ws_exec.cell(row=r, column=17).value) for r in range(8, 11))

    return {"name": name, "edrpou": edrpou, "coverage": coverage,
            "stocks": stocks, "refusals": refusals,
            "temp_contraindications": temp, "perm_contraindications": perm}


def aggregate_files(file_bytes_list, org_name, org_edrpou, report_period):
    workbooks = [load_workbook(io.BytesIO(fbytes), data_only=True) for _, fbytes in file_bytes_list]
    template_wb = load_workbook(io.BytesIO(file_bytes_list[0][1]))

    ws_out = template_wb["Виконання"]
    for row in range(8, 105): ws_out.cell(row=row, column=5).value = 0
    for row in range(8, 11):
        for col in [10,11,12,13,16,17,18]: ws_out.cell(row=row, column=col).value = 0
    for row in range(8, 14): ws_out.cell(row=row, column=20).value = 0

    for wb in workbooks:
        ws = wb["Виконання"]
        for row in range(8, 105):
            ws_out.cell(row=row, column=5).value = (ws_out.cell(row=row, column=5).value or 0) + safe_num(ws.cell(row=row, column=5).value)
        for row in range(8, 11):
            for col in [10,11,12,13,16,17]:
                ws_out.cell(row=row, column=col).value = (ws_out.cell(row=row, column=col).value or 0) + safe_num(ws.cell(row=row, column=col).value)
        for row in range(8, 14):
            ws_out.cell(row=row, column=20).value = (ws_out.cell(row=row, column=20).value or 0) + safe_num(ws.cell(row=row, column=20).value)

    for row in range(8, 11):
        ws_out.cell(row=row, column=18).value = (ws_out.cell(row=row, column=16).value or 0) + (ws_out.cell(row=row, column=17).value or 0)

    group_sums = {11:range(8,12), 23:range(12,24), 35:range(24,36), 42:range(36,43), 48:range(43,49), 61:range(49,62)}
    for sr, rows in group_sums.items():
        ws_out.cell(row=sr, column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in rows)
    ws_out.cell(row=99,  column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in [62,64,66,68,70,72,74,76,78])
    ws_out.cell(row=100, column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in [63,65,67,69,71,73,75,77,79,81])
    ws_out.cell(row=101, column=7).value = safe_num(ws_out.cell(row=80, column=5).value)
    ws_out.cell(row=102, column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in [82,84,86,88,90,92,94,96,98,100,102])
    ws_out.cell(row=103, column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in [83,85,87,89,91,93,95,97,99,101])
    ws_out.cell(row=104, column=7).value = sum(safe_num(ws_out.cell(row=r, column=5).value) for r in [103,104])

    ws_rem_out = template_wb["Залишки"]
    for row in range(11, 38):
        for col in [2,3,5,6,7,8]: ws_rem_out.cell(row=row, column=col).value = 0
    for wb in workbooks:
        ws = wb["Залишки"]
        for row in range(11, 38):
            if not ws_rem_out.cell(row=row, column=1).value: continue
            for col in [2,3,5,6,7,8]:
                ws_rem_out.cell(row=row, column=col).value = (ws_rem_out.cell(row=row, column=col).value or 0) + safe_num(ws.cell(row=row, column=col).value)
    for row in range(11, 38):
        if not ws_rem_out.cell(row=row, column=1).value: continue
        b,c,f,g,h = [safe_num(ws_rem_out.cell(row=row, column=x).value) for x in [2,3,6,7,8]]
        ws_rem_out.cell(row=row, column=4).value = b+c+g+h-f

    ws_plan_out = template_wb["План"]
    for row in range(11, 47): ws_plan_out.cell(row=row, column=6).value = 0
    for wb in workbooks:
        ws = wb["План"]
        for row in range(11, 47):
            ws_plan_out.cell(row=row, column=6).value = (ws_plan_out.cell(row=row, column=6).value or 0) + safe_num(ws.cell(row=row, column=6).value)

    ws_zvit_out = template_wb["Зведений звіт"]
    for row in range(11, 120):
        for col in [3,4,5]: ws_zvit_out.cell(row=row, column=col).value = 0
    for wb in workbooks:
        ws = wb["Зведений звіт"]
        for row in range(11, 120):
            for col in [3,4,5]:
                ws_zvit_out.cell(row=row, column=col).value = (ws_zvit_out.cell(row=row, column=col).value or 0) + safe_num(ws.cell(row=row, column=col).value)
    for row in range(11, 120):
        c_val = ws_zvit_out.cell(row=row, column=3).value
        e_val = ws_zvit_out.cell(row=row, column=5).value
        ws_zvit_out.cell(row=row, column=6).value = (
            round(e_val / c_val * 100, 1)
            if c_val and isinstance(c_val,(int,float)) and c_val > 0 and isinstance(e_val,(int,float))
            else None)

    template_wb["План"]["D8"].value          = org_name
    template_wb["План"]["E8"].value          = org_edrpou
    template_wb["Виконання"]["C4"].value     = org_name
    template_wb["Виконання"]["F4"].value     = org_edrpou
    template_wb["Виконання"]["F6"].value     = report_period
    template_wb["Залишки"]["A4"].value       = org_name
    template_wb["Залишки"]["D4"].value       = org_edrpou
    template_wb["Залишки"]["D6"].value       = report_period
    template_wb["Зведений звіт"]["A3"].value = org_name
    template_wb["Зведений звіт"]["D3"].value = org_edrpou
    template_wb["Зведений звіт"]["D5"].value = report_period

    out = io.BytesIO()
    template_wb.save(out)
    out.seek(0)
    return out.getvalue()


def generate_level1_file(good_results, org_name, org_edrpou, report_period):
    """Генерує Level1 файл (плоскі таблиці) для подачі на національний рівень."""
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Заголовки аркушів
    H_ZVEDENY = [
        "Вакцина", "Вік",
        "Кількість осіб, що підлягали щепленню у звітному році",
        "Кількість осіб, яким проведено щеплення за звітний місяць",
        "Назва закладу", "код ЄДРПОУ", "Звітний період",
    ]
    H_ARKUSH2 = [
        "Назва закладу", "код ЄДРПОУ", "Вакцина", "Вік", "Кількість щеплень", "Звітний період",
        "Народилось дітей в пологому стаціонарі у звітному місяці",
        "з них підлягали на вакцинацію проти гепатиту В у першу добу життя",
        "народилося за 7 місяців до звітного періоду",
        "з них отримали КДП 3 до 6 місяців 29 днів",
        "антиген та вік", "Кількість підлягаючих щепленню дітей за звітній місяць",
        "Тимчасові", "Постійні", "ВСЬОГО",
        "ВІДМОВИ (кількість дітей до 2 років з відмовами від профілактичних щеплень, стан",
        "кількість відмов",
    ]
    H_ZALISHOK = [
        "Вакцина",
        "Залишок на початок звітного періоду (доз)", "Отримано (доз)",
        "Залишок на кінець звітного періоду (доз)", "Виконано щеплень",
        "Використано вакцини (доз)",
        "Закуплено за кошти місцевого бюджету (доз)",
        "Закуплено з інших джерел фінансування (доз)",
        "Назва закладу", "код ЄДРПОУ", "Звітний період",
    ]
    H_PLANUV = [
        "Назва закладу", "код ЄДРПОУ", "РІК",
        "Інфекційна хвороба проти  якої планується проведення профілактичного щеплення",
        "Вік", "Кількість осіб, які підлягають профілактичному щепленню",
        "Примітка", "Вакцинація",
    ]
    H_ARKUSH3 = [
        "Назва закладу", "код ЄДРПОУ", "Звітний період",
        "Народилось дітей в пологому стаціонарі у звітному місяці",
        "з них підлягали на вакцинацію проти гепатиту В у першу добу життя",
        "народилося за 7 місяців до звітного періоду",
        "з них отримали КДП 3 до 6 місяців 29 днів",
    ]
    H_ARKUSH4 = [
        "Назва закладу", "код ЄДРПОУ", "Звітний період", "антиген та вік",
        "Кількість підлягаючих щепленню дітей за звітній місяць",
        "Тимчасові", "Постійні", "ВСЬОГО",
    ]
    H_ARKUSH5 = [
        "ВІДМОВИ (кількість дітей до 2 років з відмовами від профілактичних щеплень, стан",
        "кількість відмов", "Звітний період", "Назва закладу", "код ЄДРПОУ",
    ]

    ws_zv  = wb_out.create_sheet("зведений")
    ws_a2  = wb_out.create_sheet("Аркуш2")
    ws_zal = wb_out.create_sheet("Залишок")
    ws_pl  = wb_out.create_sheet("Планування")
    ws_a3  = wb_out.create_sheet("Аркуш3")
    ws_a4  = wb_out.create_sheet("Аркуш4")
    ws_a5  = wb_out.create_sheet("Аркуш5")

    for ws, headers in [
        (ws_zv, H_ZVEDENY), (ws_a2, H_ARKUSH2), (ws_zal, H_ZALISHOK),
        (ws_pl, H_PLANUV),  (ws_a3, H_ARKUSH3), (ws_a4, H_ARKUSH4), (ws_a5, H_ARKUSH5),
    ]:
        ws.append(headers)

    year = report_period.year if hasattr(report_period, "year") else 2026

    for r in good_results:
        fbytes     = r["_bytes"]
        zoz_name   = r["name"]
        zoz_edrpou = r["edrpou"]
        wb = load_workbook(io.BytesIO(fbytes), data_only=True)
        ws_exec = wb["Виконання"]
        ws_rem  = wb["Залишки"]
        ws_zvit = wb["Зведений звіт"]
        ws_plan = wb["План"]

        # ── "зведений": рядки з Зведений звіт ──────────────────────
        for row in range(11, 120):
            vac = ws_zvit.cell(row=row, column=1).value
            age = ws_zvit.cell(row=row, column=2).value
            pl  = ws_zvit.cell(row=row, column=3).value
            mon = ws_zvit.cell(row=row, column=4).value
            if vac is None and mon is None:
                continue
            ws_zv.append([vac, age, pl, mon, zoz_name, zoz_edrpou, report_period])

        # ── "Аркуш2": рядки з Виконання (рядки 8-104) ───────────────
        # Дані народжуваності — лише з рядку 8
        born8  = ws_exec.cell(row=8, column=10).value
        hepv18 = ws_exec.cell(row=8, column=11).value
        born78 = ws_exec.cell(row=8, column=12).value
        kdp38  = ws_exec.cell(row=8, column=13).value

        for row in range(8, 105):
            vac = ws_exec.cell(row=row, column=3).value
            age = ws_exec.cell(row=row, column=4).value
            cnt = ws_exec.cell(row=row, column=5).value
            # народжуваність — тільки у першому рядку
            born  = born8  if row == 8 else None
            hepv1 = hepv18 if row == 8 else None
            born7 = born78 if row == 8 else None
            kdp3  = kdp38  if row == 8 else None
            # протипокази КДП — рядки 8-10
            kdp_lbl = ws_exec.cell(row=row, column=14).value if row <= 10 else None
            kdp_cnt = ws_exec.cell(row=row, column=15).value if row <= 10 else None
            temp_ci = ws_exec.cell(row=row, column=16).value if row <= 10 else None
            perm_ci = ws_exec.cell(row=row, column=17).value if row <= 10 else None
            all_ci  = ws_exec.cell(row=row, column=18).value if row <= 10 else None
            # відмови — рядки 8-13
            ref_dis = ws_exec.cell(row=row, column=19).value if row <= 13 else None
            ref_cnt = ws_exec.cell(row=row, column=20).value if row <= 13 else None
            ws_a2.append([
                zoz_name, zoz_edrpou, vac, age, cnt, report_period,
                born, hepv1, born7, kdp3,
                kdp_lbl, kdp_cnt, temp_ci, perm_ci, all_ci,
                ref_dis, ref_cnt,
            ])

        # ── "Аркуш3": народжуваність (1 рядок на ЗОЗ) ───────────────
        ws_a3.append([
            zoz_name, zoz_edrpou, report_period,
            born8, hepv18, born78, kdp38,
        ])

        # ── "Аркуш4": протипокази КДП (3 рядки на ЗОЗ) ──────────────
        for row in range(8, 11):
            ws_a4.append([
                zoz_name, zoz_edrpou, report_period,
                ws_exec.cell(row=row, column=14).value,
                ws_exec.cell(row=row, column=15).value,
                ws_exec.cell(row=row, column=16).value,
                ws_exec.cell(row=row, column=17).value,
                ws_exec.cell(row=row, column=18).value,
            ])

        # ── "Аркуш5": відмови (6 рядків на ЗОЗ) ─────────────────────
        for row in range(8, 14):
            ws_a5.append([
                ws_exec.cell(row=row, column=19).value,
                ws_exec.cell(row=row, column=20).value,
                report_period, zoz_name, zoz_edrpou,
            ])

        # ── "Залишок": рядки з Залишки ───────────────────────────────
        for row in range(11, 38):
            vac = ws_rem.cell(row=row, column=1).value
            if not vac:
                continue
            ws_zal.append([
                vac,
                ws_rem.cell(row=row, column=2).value,
                ws_rem.cell(row=row, column=3).value,
                ws_rem.cell(row=row, column=4).value,
                ws_rem.cell(row=row, column=5).value,
                ws_rem.cell(row=row, column=6).value,
                ws_rem.cell(row=row, column=7).value,
                ws_rem.cell(row=row, column=8).value,
                zoz_name, zoz_edrpou, report_period,
            ])

        # ── "Планування": рядки з План ───────────────────────────────
        for row in range(11, 47):
            noz = ws_plan.cell(row=row, column=4).value
            if not noz:
                continue
            ws_pl.append([
                zoz_name, zoz_edrpou, year,
                noz,
                ws_plan.cell(row=row, column=5).value,
                ws_plan.cell(row=row, column=6).value,
                ws_plan.cell(row=row, column=7).value,
                ws_plan.cell(row=row, column=8).value,
            ])

    # Стиль заголовків
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for ws in [ws_zv, ws_a2, ws_zal, ws_pl, ws_a3, ws_a4, ws_a5]:
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    out = io.BytesIO()
    wb_out.save(out)
    out.seek(0)
    return out.getvalue()


def build_coverage_excel(facility_data_list, selected_facilities, selected_indicators):
    all_labels = []
    seen = set()
    for fd in facility_data_list:
        for item in fd["coverage"]:
            if item["label"] not in seen:
                all_labels.append(item["label"])
                seen.add(item["label"])
    if selected_indicators:
        all_labels = [l for l in all_labels if l in selected_indicators]
    facilities = [fd for fd in facility_data_list if fd["name"] in selected_facilities]

    wb = Workbook()
    ws = wb.active
    ws.title = "Охоплення"

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    subhead_fill = PatternFill("solid", fgColor="2E75B6")
    subhead_font = Font(color="FFFFFF", bold=True, size=10)
    green_fill   = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill  = PatternFill("solid", fgColor="FFEB9C")
    red_fill     = PatternFill("solid", fgColor="FFC7CE")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    end_col = get_col_letter(1 + len(all_labels))
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = f"Таблиця охоплення щепленнями — {st.session_state.get('report_label', '')}"
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["A1"].alignment = center

    ws["A2"] = "Заклад"
    ws["A2"].fill = subhead_fill; ws["A2"].font = subhead_font
    ws["A2"].alignment = center; ws["A2"].border = border
    ws.column_dimensions["A"].width = 35

    for ci, label in enumerate(all_labels, start=2):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = subhead_fill; cell.font = subhead_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_col_letter(ci)].width = 16
    ws.row_dimensions[2].height = 60

    for ri, fd in enumerate(facilities, start=3):
        ca = ws.cell(row=ri, column=1, value=fd["name"])
        ca.alignment = Alignment(vertical="center", wrap_text=True)
        ca.border = border
        for ci, label in enumerate(all_labels, start=2):
            pct = next((item["pct"] for item in fd["coverage"] if item["label"] == label), None)
            cell = ws.cell(row=ri, column=ci)
            if pct is not None:
                cell.value = pct / 100; cell.number_format = "0.0%"
                cell.fill = green_fill if pct >= 95 else (yellow_fill if pct >= 85 else red_fill)
            else:
                cell.value = "—"
            cell.alignment = center; cell.border = border

    lr = len(facilities) + 4
    ws.cell(row=lr,   column=1, value="Легенда:").font = Font(bold=True)
    ws.cell(row=lr+1, column=1, value="≥ 95% — виконання плану").fill  = green_fill
    ws.cell(row=lr+2, column=1, value="85–95% — потребує уваги").fill  = yellow_fill
    ws.cell(row=lr+3, column=1, value="< 85% — критично").fill         = red_fill

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ─────────────────────────────────────────────
# ІНТЕРФЕЙС
# ─────────────────────────────────────────────

# ── КРОК 1: Реквізити ────────────────────────────────────────────────
st.header("🏥 Крок 1 — Реквізити організації що формує зведений звіт")

with st.form("org_form"):
    col1, col2 = st.columns([3, 1])
    with col1:
        org_name = st.text_input("Назва закладу", placeholder="Наприклад: Тернопільський обласний ЦКПХ")
    with col2:
        org_edrpou = st.text_input("Код ЄДРПОУ", placeholder="12345678")
    col3, col4, col5 = st.columns(3)
    with col3:
        report_year = st.selectbox("Звітний рік", options=list(range(2024, 2031)), index=2)
    with col4:
        month_names = {1:"Січень",2:"Лютий",3:"Березень",4:"Квітень",5:"Травень",6:"Червень",
                       7:"Липень",8:"Серпень",9:"Вересень",10:"Жовтень",11:"Листопад",12:"Грудень"}
        report_month = st.selectbox("Звітний місяць", options=list(month_names.keys()),
                                    format_func=lambda x: month_names[x], index=1)
    with col5:
        expected_count = st.number_input("Кількість ЗОЗ що мають подати звіт",
                                         min_value=1, max_value=500, value=None,
                                         step=1, placeholder="Введіть кількість...")
    submitted = st.form_submit_button("✅ Підтвердити реквізити", type="primary", use_container_width=True)

if submitted:
    errs = []
    if not org_name.strip():       errs.append("Введіть назву закладу")
    if not org_edrpou.strip():     errs.append("Введіть код ЄДРПОУ")
    elif not org_edrpou.strip().isdigit(): errs.append("Код ЄДРПОУ повинен містити тільки цифри")
    if expected_count is None:     errs.append("Введіть кількість ЗОЗ що мають подати звіт")
    if errs:
        for e in errs: st.error(f"❌ {e}")
    else:
        st.session_state.update({
            "org_name":       org_name.strip(),
            "org_edrpou":     org_edrpou.strip(),
            "report_period":  datetime(report_year, report_month, 1),
            "report_label":   f"{month_names[report_month]} {report_year}",
            "expected_count": int(expected_count),
        })

if "org_name" in st.session_state:
    st.success(
        f"✅ **{st.session_state['org_name']}**  |  "
        f"ЄДРПОУ: **{st.session_state['org_edrpou']}**  |  "
        f"Звітний період: **{st.session_state['report_label']}**  |  "
        f"Очікується ЗОЗ: **{st.session_state['expected_count']}**"
    )
    st.divider()

    # ── КРОК 2: Завантаження ─────────────────────────────────────────
    st.header("📂 Крок 2 — Завантажте файли ЗОЗ")
    st.info("💡 Щоб обрати кілька файлів одночасно — утримуйте **Ctrl** (або **Cmd** на Mac) при виборі.")

    uploaded_files = st.file_uploader("Оберіть файли Excel (.xlsx)", type=["xlsx"],
                                      accept_multiple_files=True)

    # Зберігаємо файли у session_state щоб не губились після rerun
    if uploaded_files:
        st.session_state["uploaded_files_bytes"] = [(f.name, f.read(), f.size) for f in uploaded_files]

    # Беремо файли або з поточного завантаження, або зі збереженого стану
    files_in_memory = st.session_state.get("uploaded_files_bytes", [])

    if files_in_memory:
        exp      = st.session_state["expected_count"]
        sub_cnt  = len(files_in_memory)
        miss_cnt = max(0, exp - sub_cnt)
        pct      = min(100, round(sub_cnt / exp * 100))

        st.subheader("📊 Стан подання звітів")
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("📋 Очікується", exp)
        m2.metric("📥 Подано", sub_cnt)
        m3.metric("⏳ Не подали", miss_cnt, delta=f"-{miss_cnt}" if miss_cnt else None, delta_color="inverse")
        m4.metric("✅ Виконання", f"{pct}%")
        st.progress(pct/100, text=f"Подано {sub_cnt} з {exp} закладів ({pct}%)")
        if miss_cnt == 0: st.success("🎉 Усі заклади подали звіти!")
        else: st.warning(f"⚠️ Ще не подали звіт: **{miss_cnt}** заклад(ів).")

        st.divider()
        st.header("📋 Крок 3 — Список завантажених файлів")
        st.dataframe(pd.DataFrame([{"Файл": fname, "Розмір": f"{round(fsize/1024,1)} КБ",
                                    "Статус": "⏳ Очікує перевірки"}
                                   for fname, _, fsize in files_in_memory]),
                     use_container_width=True, hide_index=True)

        col_check, col_reset = st.columns([3, 1])
        with col_reset:
            if st.button("🗑️ Очистити файли", use_container_width=True):
                st.session_state.pop("uploaded_files_bytes", None)
                st.session_state.pop("results", None)
                st.session_state.pop("corrections_log", None)
                st.rerun()
        with col_check:
            st.header("🔍 Крок 4 — Перевірка файлів")
        if st.button("▶️ Запустити перевірку", type="primary", use_container_width=True):
            results = []
            pb = st.progress(0, text="Перевірка виконується...")
            for i, (fname, fbytes, _) in enumerate(files_in_memory):
                r = validate_file(fbytes, fname)
                r["_bytes"] = fbytes
                results.append(r)
                pb.progress((i+1)/len(files_in_memory), text=f"Перевірено {i+1} з {len(files_in_memory)}: {fname}")
            st.session_state["results"] = results
            st.session_state["corrections_log"] = {}
            st.rerun()


# ─── Результати + Виправлення ─────────────────────────────────────────
if "results" in st.session_state and "org_name" in st.session_state:
    results  = st.session_state["results"]
    exp      = st.session_state["expected_count"]
    corr_log = st.session_state.get("corrections_log", {})

    ok   = [r for r in results if r["status"] == "🟢 OK"]
    warn = [r for r in results if r["status"] == "🟡 Попередження"]
    bad  = [r for r in results if r["status"] == "🔴 Помилки"]
    good = [r for r in results if r["status"] in ("🟢 OK","🟡 Попередження")]
    miss = max(0, exp - len(results))

    st.subheader("📊 Підсумок після перевірки")
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("📋 Очікується", exp)
    c2.metric("📥 Подано", len(results))
    c3.metric("🟢 Без помилок", len(ok))
    c4.metric("🟡 З попередженнями", len(warn))
    c5.metric("🔴 З помилками", len(bad))
    if miss > 0: st.warning(f"⚠️ Ще не подали звіт: **{miss}** заклад(ів) з {exp}")

    periods = list({r["period"] for r in results if r["period"] != "—"})
    if len(periods) > 1: st.warning(f"⚠️ Різні звітні періоди: {', '.join(periods)}")
    edrpous = [r["edrpou"] for r in results if r["edrpou"] != "—"]
    dups = {e for e in edrpous if edrpous.count(e) > 1}
    if dups: st.error(f"❌ Дублікати ЄДРПОУ: {', '.join(dups)}")

    # ── Детальні результати + форми виправлення ───────────────────────
    st.subheader("📋 Детальні результати по файлах")

    for idx, r in enumerate(results):
        fname = r["file"]
        is_corrected = fname in corr_log

        with st.expander(
            f"{r['status']}  |  {fname}  |  {r['name']}  |  ЄДРПОУ: {r['edrpou']}"
            + ("  ✏️ *виправлено*" if is_corrected else "")
        ):
            if r["errors"]:
                st.error("**Критичні помилки:**")
                for e in r["errors"]: st.write(f"• {e}")

            if r["warnings"]:
                st.warning("**Попередження:**")
                for w in r["warnings"]: st.write(f"• {w}")

            if not r["errors"] and not r["warnings"]:
                st.success("Файл пройшов усі перевірки!")

            # ── Форма виправлення (тільки якщо є помилки) ────────────
            fixable = r.get("fixable", {})
            has_fixable = any([
                "org_name" in fixable,
                "org_edrpou" in fixable,
                "report_period" in fixable,
                "neg_values" in fixable,
                "balance_errors" in fixable,
                "contra_errors" in fixable,
            ])

            if (r["errors"] or r["warnings"]) and has_fixable:
                st.divider()
                st.markdown("#### ✏️ Виправити помилки онлайн")

                corr = {}

                # Назва закладу
                if "org_name" in fixable:
                    new_name = st.text_input(
                        "Нова назва закладу:",
                        value=fixable.get("org_name", ""),
                        key=f"fix_name_{idx}"
                    )
                    if new_name.strip(): corr["org_name"] = new_name.strip()

                # ЄДРПОУ
                if "org_edrpou" in fixable:
                    new_edrpou = st.text_input(
                        f"Виправити ЄДРПОУ (поточний: {fixable.get('org_edrpou','—')}):",
                        value=fixable.get("org_edrpou",""),
                        key=f"fix_edrpou_{idx}"
                    )
                    if new_edrpou.strip(): corr["org_edrpou"] = new_edrpou.strip()

                # Звітний період
                if "report_period" in fixable:
                    col_y, col_m = st.columns(2)
                    with col_y:
                        fix_year = st.selectbox("Рік звіту:", range(2024,2031),
                                                index=2, key=f"fix_year_{idx}")
                    with col_m:
                        fix_month = st.selectbox("Місяць звіту:",
                                                 list(month_names.keys()),
                                                 format_func=lambda x: month_names[x],
                                                 index=1, key=f"fix_month_{idx}")
                    corr["report_period"] = datetime(fix_year, fix_month, 1)

                # Від'ємні значення
                if "neg_values" in fixable:
                    st.markdown("**Від'ємні значення щеплень — виправити на:**")
                    neg_corr = {}
                    for row, info in fixable["neg_values"].items():
                        new_val = st.number_input(
                            f"{info['vaccine']} / {info['age']} (зараз: {info['value']})",
                            min_value=0, value=0, step=1,
                            key=f"fix_neg_{idx}_{row}"
                        )
                        neg_corr[str(row)] = new_val
                    corr["neg_values"] = neg_corr

                # Баланс залишків
                if "balance_errors" in fixable:
                    st.markdown("**Помилки балансу залишків — прийняти розрахункове значення:**")
                    bal_accept = {}
                    for row, info in fixable["balance_errors"].items():
                        accept = st.checkbox(
                            f"{info['vaccine']}: замінити {info['current']} → {info['expected']}",
                            value=True, key=f"fix_bal_{idx}_{row}"
                        )
                        if accept: bal_accept[str(row)] = True
                    corr["balance_accept"] = bal_accept

                # Протипокази
                if "contra_errors" in fixable:
                    st.markdown("**Протипокази — виправити значення:**")
                    contra_fix = {}
                    for row, info in fixable["contra_errors"].items():
                        c1, c2 = st.columns(2)
                        with c1:
                            t = st.number_input(f"Рядок {row} — Тимчасові",
                                                min_value=0, value=int(info["temp"]),
                                                key=f"fix_ct_{idx}_{row}")
                        with c2:
                            p = st.number_input(f"Рядок {row} — Постійні",
                                                min_value=0, value=int(info["perm"]),
                                                key=f"fix_cp_{idx}_{row}")
                        contra_fix[str(row)] = {"temp": t, "perm": p}
                        st.caption(f"ВСЬОГО буде: {t + p}")
                    corr["contra_fix"] = contra_fix

                # Кнопка "Застосувати"
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("💾 Застосувати виправлення", key=f"apply_{idx}",
                                 type="primary", use_container_width=True):
                        try:
                            new_bytes = apply_corrections(r["_bytes"], corr)
                            # Перевіряємо знову
                            new_result = validate_file(new_bytes, fname)
                            # Оновлюємо results
                            results[idx]["_bytes"]   = new_bytes
                            results[idx]["status"]   = new_result["status"]
                            results[idx]["errors"]   = new_result["errors"]
                            results[idx]["warnings"] = new_result["warnings"]
                            results[idx]["fixable"]  = new_result["fixable"]
                            results[idx]["name"]     = new_result["name"]
                            results[idx]["edrpou"]   = new_result["edrpou"]

                            # Логуємо виправлення
                            fix_log = []
                            if corr.get("org_name"):     fix_log.append(f"Назва закладу → {corr['org_name']}")
                            if corr.get("org_edrpou"):   fix_log.append(f"ЄДРПОУ → {corr['org_edrpou']}")
                            if corr.get("report_period"): fix_log.append(f"Звітний період → {corr['report_period'].strftime('%m.%Y')}")
                            if corr.get("neg_values"):   fix_log.append(f"Виправлено від'ємних значень: {len(corr['neg_values'])}")
                            if corr.get("balance_accept"): fix_log.append(f"Виправлено балансів залишків: {len(corr['balance_accept'])}")
                            if corr.get("contra_fix"):   fix_log.append(f"Виправлено протипоказів: {len(corr['contra_fix'])}")

                            corr_log[fname] = fix_log
                            st.session_state["results"]          = results
                            st.session_state["corrections_log"]  = corr_log
                            st.success(f"✅ Виправлення застосовано! Новий статус: {new_result['status']}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Помилка: {e}")

                with col_btn2:
                    # Завантажити виправлений файл
                    if fname in corr_log:
                        st.download_button(
                            "⬇️ Завантажити виправлений файл",
                            data=results[idx]["_bytes"],
                            file_name=f"виправлений_{fname}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key=f"dl_fix_{idx}"
                        )

    # ── PDF-звіт ─────────────────────────────────────────────────────
    st.divider()
    col_pdf1, col_pdf2 = st.columns([1, 2])
    with col_pdf1:
        if st.button("📄 Сформувати PDF-звіт про перевірку", use_container_width=True):
            with st.spinner("Формування PDF..."):
                try:
                    pdf_bytes = generate_pdf_report(
                        results, corr_log,
                        st.session_state["org_name"],
                        st.session_state["report_label"]
                    )
                    period_str = st.session_state["report_label"].replace(" ","_")
                    st.session_state["pdf_bytes"] = pdf_bytes
                    st.session_state["pdf_name"]  = f"Звіт_перевірки_{period_str}.pdf"
                except Exception as e:
                    st.error(f"❌ Помилка формування PDF: {e}")
    with col_pdf2:
        if "pdf_bytes" in st.session_state:
            st.download_button(
                "⬇️ Завантажити PDF-звіт",
                data=st.session_state["pdf_bytes"],
                file_name=st.session_state["pdf_name"],
                mime="application/pdf",
                type="primary", use_container_width=True
            )

    # ── КРОК 5: Зведення ─────────────────────────────────────────────
    good = [r for r in results if r["status"] in ("🟢 OK","🟡 Попередження")]
    if good:
        st.divider()
        st.header("📊 Крок 5 — Результати")
        if miss > 0:
            st.warning(f"⚠️ Увага: **{miss}** заклад(ів) ще не подали звіт. Зведення буде неповним.")
        if bad:
            with st.expander("🔴 Виключені файли (критичні помилки)"):
                for r in bad: st.write(f"• {r['file']} — {r['name']}")

        facility_data_list = []
        for r in good:
            try:
                fd = extract_facility_data(r["_bytes"], r["name"], r["edrpou"])
                facility_data_list.append(fd)
            except Exception: pass

        tab_main, tab_coverage, tab_dash = st.tabs([
            "📥 Зведений файл", "📊 Таблиця охоплення", "📈 Дашборди"
        ])

        # ── Вкладка 1: Зведений файл ─────────────────────────────────
        with tab_main:
            st.markdown(f"До зведення включено **{len(good)}** файл(ів) з **{exp}** очікуваних.")
            if corr_log:
                st.info(f"ℹ️ {len(corr_log)} файл(ів) були виправлені онлайн — використовуються виправлені версії.")

            col_agg, col_lvl = st.columns(2)

            with col_agg:
                if st.button("⚙️ Створити зведений файл", type="primary", use_container_width=True):
                    with st.spinner("⏳ Зведення виконується..."):
                        try:
                            file_bytes_list = [(r["file"], r["_bytes"]) for r in good]
                            result_bytes = aggregate_files(
                                file_bytes_list,
                                org_name      = st.session_state["org_name"],
                                org_edrpou    = st.session_state["org_edrpou"],
                                report_period = st.session_state["report_period"]
                            )
                            period_str = st.session_state["report_label"].replace(" ","_")
                            st.session_state["agg_bytes"] = result_bytes
                            st.session_state["agg_name"]  = f"Зведений_звіт_{period_str}.xlsx"
                            st.success(f"✅ Успішно зведено {len(good)} файлів!")
                        except Exception as e:
                            st.error(f"❌ Помилка: {e}")
            if "agg_bytes" in st.session_state:
                st.download_button(
                    "⬇️ Завантажити зведений файл", data=st.session_state["agg_bytes"],
                    file_name=st.session_state["agg_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary", key="dl_agg"
                )

            st.divider()

            with col_lvl:
                if st.button("🏛️ Сформувати Level 1 файл", use_container_width=True,
                             help="Плоскі таблиці для подачі на національний рівень (МОЗ/УЦКПХ)"):
                    with st.spinner("⏳ Формування Level 1..."):
                        try:
                            lvl1_bytes = generate_level1_file(
                                good,
                                org_name      = st.session_state["org_name"],
                                org_edrpou    = st.session_state["org_edrpou"],
                                report_period = st.session_state["report_period"],
                            )
                            period_str = st.session_state["report_label"].replace(" ","_")
                            st.session_state["lvl1_bytes"] = lvl1_bytes
                            st.session_state["lvl1_name"]  = f"Level1_{period_str}.xlsx"
                            st.success(f"✅ Level 1 файл сформовано ({len(good)} ЗОЗ)!")
                        except Exception as e:
                            st.error(f"❌ Помилка: {e}")
            if "lvl1_bytes" in st.session_state:
                st.download_button(
                    "⬇️ Завантажити Level 1", data=st.session_state["lvl1_bytes"],
                    file_name=st.session_state["lvl1_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_lvl1"
                )
                with st.expander("ℹ️ Що містить Level 1 файл"):
                    st.markdown("""
| Аркуш | Вміст | Рядків |
|---|---|---|
| `зведений` | Зведений звіт усіх ЗОЗ (вакцина/вік/план/місяць) | ~109 × кількість ЗОЗ |
| `Аркуш2` | Деталізоване виконання + народжуваність + протипокази + відмови | ~97 × ЗОЗ |
| `Залишок` | Залишки вакцин усіх ЗОЗ | ~27 × ЗОЗ |
| `Планування` | Плани вакцинації усіх ЗОЗ | ~36 × ЗОЗ |
| `Аркуш3` | Народжуваність і КДП (1 рядок на ЗОЗ) | = кількість ЗОЗ |
| `Аркуш4` | Протипокази КДП (3 рядки на ЗОЗ) | = 3 × ЗОЗ |
| `Аркуш5` | Відмови від щеплень (6 рядків на ЗОЗ) | = 6 × ЗОЗ |
""")

        # ── Вкладка 2: Таблиця охоплення ─────────────────────────────
        with tab_coverage:
            all_facility_names = [fd["name"] for fd in facility_data_list]
            all_labels = []
            seen = set()
            for fd in facility_data_list:
                for item in fd["coverage"]:
                    if item["label"] not in seen:
                        all_labels.append(item["label"]); seen.add(item["label"])

            col_f, col_i = st.columns(2)
            with col_f:
                st.markdown("**Заклади:**")
                sel_all_f = st.checkbox("Обрати всі заклади", value=True, key="all_fac")
                selected_facilities = all_facility_names if sel_all_f else st.multiselect(
                    "Оберіть заклади:", all_facility_names, default=all_facility_names)
            with col_i:
                st.markdown("**Показники (вакцини):**")
                sel_all_i = st.checkbox("Обрати всі показники", value=True, key="all_ind")
                selected_indicators = all_labels if sel_all_i else st.multiselect(
                    "Оберіть показники:", all_labels, default=all_labels)

            if selected_facilities and selected_indicators:
                if st.button("⚙️ Сформувати таблицю охоплення", type="primary", use_container_width=True):
                    with st.spinner("⏳ Формування..."):
                        try:
                            cov_bytes = build_coverage_excel(facility_data_list, selected_facilities, selected_indicators)
                            period_str = st.session_state["report_label"].replace(" ","_")
                            st.success("✅ Таблиця охоплення сформована!")
                            st.download_button("⬇️ Завантажити таблицю охоплення", data=cov_bytes,
                                               file_name=f"Таблиця_охоплення_{period_str}.xlsx",
                                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                               use_container_width=True, type="primary")
                        except Exception as e:
                            st.error(f"❌ Помилка: {e}")

        # ── Вкладка 3: Дашборди ───────────────────────────────────────
        with tab_dash:
            d1,d2,d3,d4 = st.tabs(["💉 Охоплення","🏆 Рейтинг ЗОЗ","🏥 Залишки вакцин","❌ Відмови та протипокази"])

            with d1:
                st.subheader("💉 Охоплення щепленнями")
                agg = {}
                for fd in facility_data_list:
                    for item in fd["coverage"]:
                        if item["label"] not in agg: agg[item["label"]] = {"plan":0,"executed":0}
                        agg[item["label"]]["plan"]     += item["plan"]
                        agg[item["label"]]["executed"] += item["executed"]
                df_agg = pd.DataFrame([{"Вакцина / Вік": l, "План": v["plan"], "Виконано": v["executed"],
                                        "%": round(v["executed"]/v["plan"]*100,1) if v["plan"]>0 else 0}
                                       for l,v in agg.items()]).sort_values("%", ascending=True)
                all_v = df_agg["Вакцина / Вік"].tolist()
                sel_v = st.multiselect("Оберіть вакцини:", all_v, default=all_v[:15] if len(all_v)>15 else all_v, key="dv")
                df_f = df_agg[df_agg["Вакцина / Вік"].isin(sel_v)]
                if not df_f.empty:
                    colors_list = ["#2ECC71" if p>=95 else "#F39C12" if p>=85 else "#E74C3C" for p in df_f["%"]]
                    fig = go.Figure(go.Bar(x=df_f["%"], y=df_f["Вакцина / Вік"], orientation="h",
                                          marker_color=colors_list, text=[f"{p}%" for p in df_f["%"]],
                                          textposition="outside",
                                          hovertemplate="<b>%{y}</b><br>%{x}%<extra></extra>"))
                    fig.add_vline(x=95, line_dash="dash", line_color="green", annotation_text="95%")
                    fig.add_vline(x=85, line_dash="dash", line_color="orange", annotation_text="85%")
                    fig.update_layout(xaxis=dict(range=[0,120]), xaxis_title="% виконання",
                                      height=max(400,len(df_f)*30), margin=dict(l=10,r=80,t=20,b=40),
                                      plot_bgcolor="white")
                    st.plotly_chart(fig, use_container_width=True)
                    lc1,lc2,lc3 = st.columns(3)
                    lc1.success("🟢 ≥ 95%"); lc2.warning("🟡 85–95%"); lc3.error("🔴 < 85%")

            with d2:
                st.subheader("🏆 Рейтинг закладів")
                all_vr = sorted({item["label"] for fd in facility_data_list for item in fd["coverage"]})
                mode = st.radio("Показник:", ["Середній % по всіх вакцинах","По обраній вакцині"], horizontal=True)
                sel_vr = st.selectbox("Оберіть вакцину:", all_vr, key="rv") if mode=="По обраній вакцині" else None
                rows_r = []
                for fd in facility_data_list:
                    if mode == "Середній % по всіх вакцинах":
                        pcts = [i["pct"] for i in fd["coverage"] if i["plan"]>0]
                        rows_r.append({"Заклад":fd["name"], "%": round(sum(pcts)/len(pcts),1) if pcts else 0})
                    else:
                        p = next((i["pct"] for i in fd["coverage"] if i["label"]==sel_vr), 0)
                        rows_r.append({"Заклад":fd["name"], "%":p})
                df_r = pd.DataFrame(rows_r).sort_values("%", ascending=False).reset_index(drop=True)
                df_r.index += 1
                colors_r = ["#2ECC71" if p>=95 else "#F39C12" if p>=85 else "#E74C3C" for p in df_r["%"]]
                fig_r = go.Figure(go.Bar(x=df_r["%"], y=df_r["Заклад"], orientation="h", marker_color=colors_r,
                                         text=[f"{p}%" for p in df_r["%"]], textposition="outside",
                                         hovertemplate="<b>%{y}</b><br>%{x}%<extra></extra>"))
                fig_r.add_vline(x=95, line_dash="dash", line_color="green")
                fig_r.add_vline(x=85, line_dash="dash", line_color="orange")
                fig_r.update_layout(xaxis=dict(range=[0,120]), yaxis=dict(autorange="reversed"),
                                     height=max(300,len(df_r)*40), margin=dict(l=10,r=80,t=20,b=40),
                                     plot_bgcolor="white")
                st.plotly_chart(fig_r, use_container_width=True)
                df_r["Місце"] = range(1,len(df_r)+1)
                df_r["Статус"] = df_r["%"].apply(lambda x: "🟢 OK" if x>=95 else ("🟡 Увага" if x>=85 else "🔴 Критично"))
                st.dataframe(df_r[["Місце","Заклад","%","Статус"]], use_container_width=True, hide_index=True)

            with d3:
                st.subheader("🏥 Залишки вакцин")
                thr = st.number_input("⚙️ Поріг критичного залишку (доз):", min_value=0, max_value=10000, value=50, step=10)
                sagg = {}
                for fd in facility_data_list:
                    for s in fd["stocks"]:
                        if s["vaccine"] not in sagg: sagg[s["vaccine"]] = {"closing":0,"used":0}
                        sagg[s["vaccine"]]["closing"] += s["closing"]
                        sagg[s["vaccine"]]["used"]    += s["used"]
                df_s = pd.DataFrame([{"Вакцина":v,"Залишок (доз)":d["closing"],"Витрачено (доз)":d["used"]}
                                      for v,d in sagg.items() if d["closing"]>0 or d["used"]>0]
                                    ).sort_values("Залишок (доз)", ascending=True)
                if not df_s.empty:
                    cs = ["#E74C3C" if v<=thr else "#3498DB" for v in df_s["Залишок (доз)"]]
                    fig_s = go.Figure(go.Bar(x=df_s["Залишок (доз)"], y=df_s["Вакцина"], orientation="h",
                                            marker_color=cs, text=df_s["Залишок (доз)"].astype(int),
                                            textposition="outside",
                                            hovertemplate="<b>%{y}</b><br>%{x} доз<extra></extra>"))
                    fig_s.add_vline(x=thr, line_dash="dash", line_color="red",
                                   annotation_text=f"Критично ({thr})", annotation_position="top right")
                    fig_s.update_layout(height=max(400,len(df_s)*30), margin=dict(l=10,r=80,t=20,b=40),
                                        plot_bgcolor="white")
                    st.plotly_chart(fig_s, use_container_width=True)
                    crit = df_s[df_s["Залишок (доз)"]<=thr]
                    if not crit.empty:
                        st.error(f"🚨 Критично мало ({thr} доз і менше):")
                        st.dataframe(crit, use_container_width=True, hide_index=True)

            with d4:
                st.subheader("❌ Відмови та протипокази")
                cr, cc = st.columns(2)
                with cr:
                    st.markdown("**Відмови від щеплень**")
                    ref_agg = {}
                    for fd in facility_data_list:
                        for rv in fd["refusals"]: ref_agg[rv["disease"]] = ref_agg.get(rv["disease"],0) + rv["count"]
                    df_ref = pd.DataFrame([{"Нозологія":k,"Відмов":v} for k,v in ref_agg.items() if v>0])
                    if not df_ref.empty:
                        fig_ref = px.bar(df_ref, x="Відмов", y="Нозологія", orientation="h",
                                         color="Відмов", color_continuous_scale=["#FFF3CD","#E74C3C"], text="Відмов")
                        fig_ref.update_traces(textposition="outside")
                        fig_ref.update_layout(height=300, showlegend=False, plot_bgcolor="white",
                                              margin=dict(l=10,r=60,t=10,b=20))
                        st.plotly_chart(fig_ref, use_container_width=True)
                        st.dataframe(df_ref.sort_values("Відмов",ascending=False), use_container_width=True, hide_index=True)
                    else: st.info("Відмов не зареєстровано")
                with cc:
                    st.markdown("**Протипокази**")
                    tt = sum(fd["temp_contraindications"] for fd in facility_data_list)
                    tp = sum(fd["perm_contraindications"] for fd in facility_data_list)
                    if tt+tp > 0:
                        fig_p = px.pie(values=[tt,tp], names=["Тимчасові","Постійні"],
                                       color_discrete_sequence=["#F39C12","#E74C3C"], hole=0.4)
                        fig_p.update_traces(textposition="inside", textinfo="percent+label+value")
                        fig_p.update_layout(height=300, margin=dict(l=10,r=10,t=10,b=10))
                        st.plotly_chart(fig_p, use_container_width=True)
                        p1,p2,p3 = st.columns(3)
                        p1.metric("Тимчасові",int(tt)); p2.metric("Постійні",int(tp)); p3.metric("Всього",int(tt+tp))
                        contra_rows = [{"Заклад":fd["name"],"Тимчасові":int(fd["temp_contraindications"]),
                                        "Постійні":int(fd["perm_contraindications"]),
                                        "Всього":int(fd["temp_contraindications"]+fd["perm_contraindications"])}
                                       for fd in facility_data_list]
                        st.dataframe(pd.DataFrame(contra_rows).sort_values("Всього",ascending=False),
                                     use_container_width=True, hide_index=True)
                    else: st.info("Протипоказів не зареєстровано")
    else:
        st.error("❌ Немає файлів придатних для зведення — виправте помилки і спробуйте знову.")
